import os
import json
import webbrowser
import time
from datetime import datetime
from dotenv import load_dotenv
from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from compare import run_duckdb_comparison, get_common_columns
import uvicorn
from pydantic import BaseModel
from typing import List, Dict, Any
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

load_dotenv()

app = FastAPI(title="Data Comparison Report")

class TabData(BaseModel):
    tab_id: str
    name: str
    rows: List[Dict[str, Any]]

class ExportRequest(BaseModel):
    id_col_name: str
    columns: List[str]
    src_only_columns: List[str]
    dstn_only_columns: List[str]
    tabs: List[TabData]

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

def get_dsns():
    src_dsn = os.environ.get('SOURCE_DSN')
    dstn_dsn = os.environ.get('DESTINATION_DSN')
    if not src_dsn or not dstn_dsn:
        raise ValueError("SOURCE_DSN and DESTINATION_DSN must be set in the environment variables")
    return src_dsn, dstn_dsn

@app.get("/", response_class=HTMLResponse)
async def setup_page(request: Request):
    try:
        src_dsn, dstn_dsn = get_dsns()
        common_cols = get_common_columns(src_dsn, dstn_dsn)
        return templates.TemplateResponse(
            request=request,
            name="setup_template.html", 
            context={"request": request, "dsn1": src_dsn, "dsn2": dstn_dsn, "common_cols": common_cols}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/generate_report", response_class=HTMLResponse)
async def generate_report_route(request: Request, id_cols: list[str] = Form(...)):
    try:
        start_time = time.perf_counter()
        
        src_dsn, dstn_dsn = get_dsns()
        
        # Make sure user picked at least one column
        if not id_cols:
            raise HTTPException(status_code=400, detail="You must select at least one comparison key.")
            
        results = run_duckdb_comparison(src_dsn, dstn_dsn, id_cols)
        
        end_time = time.perf_counter()
        execution_time = round(end_time - start_time, 2)
        
        report_data_dict = {
            'totalIntersection': results['total_intersection'],
            'srcOnlyCount': results['src_only_count'],
            'dstnOnlyCount': results['dstn_only_count'],
            'totalMatched': results['total_intersection'] - results['total_mismatches'],
            'totalMismatched': results['total_mismatches'],
            'totalCleanRows': results['total_intersection'] - results['total_rows_with_nan'],
            'totalMissingRows': results['total_rows_with_nan'],
            'columnMismatchLabels': list(results['mismatches_by_column'].keys()),
            'columnMismatchValues': list(results['mismatches_by_column'].values()),
            'completenessLabels': results['data_columns'],
            'completenessCombined': results['completeness_combined'],
            'dataColumns': results['data_columns'],
            'srcOnlyColumns': results['src_only_columns'],
            'dstnOnlyColumns': results['dstn_only_columns'],
            'id_col_name': results['id_col_name'],
            'tableData': {
                'sub-all': results['comparison_data'],
                'sub-comparable': [r for r in results['comparison_data'] if r.get('is_comparable')],
                'sub-non-comparable-records': [r for r in results['comparison_data'] if r.get('is_src_only') or r.get('is_dstn_only')],
                'sub-non-comparable-columns': results['comparison_data'],
                'sub-matched': results['matched_data'],
                'sub-mismatched': results['mismatched_data'],
                'sub-missing': results['missing_value_rows']
            }
        }
        
        report_json_str = json.dumps(report_data_dict).replace("\\", "\\\\").replace("'", "\\'")

        # Render template context
        context = {
            "request": request,
            "id_col_name": results['id_col_name'],
            "total_intersection": results['total_intersection'],
            "total_mismatches": results['total_mismatches'],
            "total_rows_with_nan": results['total_rows_with_nan'],
            "total_columns": len(results['data_columns']),
            "data_columns": results['data_columns'],
            "comparison_data": results['comparison_data'],
            "mismatched_data": results['mismatched_data'],
            "missing_value_rows": results['missing_value_rows'],
            "matched_data": results['matched_data'],
            "column_mismatch_labels": list(results['mismatches_by_column'].keys()),
            "column_mismatch_values": list(results['mismatches_by_column'].values()),
            "completeness_labels": results['data_columns'],
            "completeness_combined": results['completeness_combined'],
            "report_json": report_json_str,
            "column_stats": results['column_stats'],
            "dsn1": src_dsn,
            "dsn2": dstn_dsn,
            "tb1": results.get('tb1'),
            "tb2": results.get('tb2'),
            "file1": results.get('file1'),
            "file2": results.get('file2'),
            "cols1_len": results.get('cols1_len'),
            "cols2_len": results.get('cols2_len'),
            "count1": results.get('count1'),
            "count2": results.get('count2'),
            "src_only_columns": results.get('src_only_columns', []),
            "dstn_only_columns": results.get('dstn_only_columns', []),
            "src_only_count": results.get('src_only_count', 0),
            "dstn_only_count": results.get('dstn_only_count', 0),
            "generation_date": datetime.now().strftime("%B %d, %Y at %I:%M %p"),
            "execution_time": execution_time
        }
        
        # Render HTML string
        template = templates.get_template('report_template.html')
        html_out = template.render(context)
        
        # Save to static directory
        output_path = os.path.join('static', 'report.html')
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_out)
        
        # Open in browser automatically
        webbrowser.open('http://127.0.0.1:8000/static/report.html')
        

        success_html = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Report Generated</title>
            <link rel="stylesheet" href="/static/style.css">
        </head>
        <body>
            <div class="mesh-bg">
                <div class="mesh-blob blob-1"></div>
                <div class="mesh-blob blob-2"></div>
                <div class="mesh-blob blob-3"></div>
            </div>
            <div class="app-wrapper" style="align-items: center; justify-content: center; min-height: 80vh;">
                <div class="glass-panel" style="padding: 3rem; text-align: center; max-width: 600px;">
                    <h2 style="color: var(--color-success); margin-bottom: 1rem;">Report Generated Successfully!</h2>
                    <p style="color: var(--color-text-secondary); margin-bottom: 2rem;">
                        You can close this window safely.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        return HTMLResponse(content=success_html)

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export_excel")
async def export_excel(req: ExportRequest):
    wb = Workbook()
    wb.remove(wb.active)
    
    fill_mismatch = PatternFill(start_color="FFB3B0", end_color="FFB3B0", fill_type="solid")
    fill_missing = PatternFill(start_color="FFF0D9", end_color="FFF0D9", fill_type="solid")
    fill_src_only = PatternFill(start_color="E2F0FF", end_color="E2F0FF", fill_type="solid")
    fill_dstn_only = PatternFill(start_color="F5E8FE", end_color="F5E8FE", fill_type="solid")
    
    fill_src_col = PatternFill(start_color="F7FBFF", end_color="F7FBFF", fill_type="solid")
    fill_dstn_col = PatternFill(start_color="FCF7FF", end_color="FCF7FF", fill_type="solid")
    font_src_col = Font(bold=True, color="0A84FF")
    font_dstn_col = Font(bold=True, color="BF5AF2")
    
    for tab in req.tabs:
        ws = wb.create_sheet(title=tab.name)
        
        if tab.tab_id == 'sub-non-comparable-columns':
            headers = [f"{req.id_col_name} (Source)"] + req.src_only_columns + [""] + [f"{req.id_col_name} (Target)"] + req.dstn_only_columns
            ws.append(headers)
            
            for c_idx in range(2, 2 + len(req.src_only_columns)):
                ws.cell(row=1, column=c_idx).font = font_src_col
                ws.cell(row=1, column=c_idx).fill = fill_src_col
            for c_idx in range(4 + len(req.src_only_columns), 4 + len(req.src_only_columns) + len(req.dstn_only_columns)):
                ws.cell(row=1, column=c_idx).font = font_dstn_col
                ws.cell(row=1, column=c_idx).fill = fill_dstn_col
                
            for row in tab.rows:
                row_data = [row.get('id_display', '')]
                for col in req.src_only_columns:
                    val = row.get(f"{col}_src_only", "")
                    row_data.append(val if val is not None else "")
                row_data.append("")
                row_data.append(row.get('id_display', ''))
                for col in req.dstn_only_columns:
                    val = row.get(f"{col}_dstn_only", "")
                    row_data.append(val if val is not None else "")
                ws.append(row_data)
                
                r_idx = ws.max_row
                for c_idx in range(2, 2 + len(req.src_only_columns)):
                    ws.cell(row=r_idx, column=c_idx).fill = fill_src_col
                for c_idx in range(4 + len(req.src_only_columns), 4 + len(req.src_only_columns) + len(req.dstn_only_columns)):
                    ws.cell(row=r_idx, column=c_idx).fill = fill_dstn_col
        else:
            col_idx = 1
            ws.cell(row=1, column=col_idx, value=req.id_col_name)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            col_idx += 1
            
            for col in req.columns:
                ws.cell(row=1, column=col_idx, value=col)
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
                ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
                ws.cell(row=1, column=col_idx).font = Font(bold=True)
                col_idx += 2
                
            if tab.tab_id in ('sub-all', 'sub-non-comparable-src'):
                for col in req.src_only_columns:
                    ws.cell(row=1, column=col_idx, value=col)
                    ws.cell(row=1, column=col_idx).font = font_src_col
                    ws.cell(row=1, column=col_idx).fill = fill_src_col
                    col_idx += 1
            
            if tab.tab_id in ('sub-all', 'sub-non-comparable-dstn'):
                for col in req.dstn_only_columns:
                    ws.cell(row=1, column=col_idx, value=col)
                    ws.cell(row=1, column=col_idx).font = font_dstn_col
                    ws.cell(row=1, column=col_idx).fill = fill_dstn_col
                    col_idx += 1
            
            r = 2
            for row in tab.rows:
                c = 1
                ws.cell(row=r, column=c, value=row.get('id_display', ''))
                c += 1
                
                for col in req.columns:
                    is_mismatch = row.get(f"{col}_mismatch", False)
                    is_nan = row.get(f"{col}_nan", False)
                    val1 = row.get(f"{col}_1")
                    val2 = row.get(f"{col}_2")
                    
                    if val1 is None: val1 = "null"
                    if val2 is None: val2 = "null"
                    
                    if row.get('is_src_only'):
                        ws.cell(row=r, column=c, value=val1)
                        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
                        ws.cell(row=r, column=c).fill = fill_src_only
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    elif row.get('is_dstn_only'):
                        ws.cell(row=r, column=c, value=val2)
                        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
                        ws.cell(row=r, column=c).fill = fill_dstn_only
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    elif is_mismatch or is_nan:
                        c1 = ws.cell(row=r, column=c, value=val1)
                        c1.fill = fill_src_only
                        c1.alignment = Alignment(horizontal='center')
                        
                        c2 = ws.cell(row=r, column=c+1, value=val2)
                        c2.fill = fill_dstn_only
                        c2.alignment = Alignment(horizontal='center')
                    else:
                        ws.cell(row=r, column=c, value=val1)
                        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    
                    c += 2
                
                if tab.tab_id in ('sub-all', 'sub-non-comparable-src'):
                    for col in req.src_only_columns:
                        val = row.get(f"{col}_src_only", "")
                        cell = ws.cell(row=r, column=c, value=val if val is not None else "")
                        if row.get('is_src_only'):
                            cell.fill = fill_src_only
                        else:
                            cell.fill = fill_src_col
                        c += 1
                
                if tab.tab_id in ('sub-all', 'sub-non-comparable-dstn'):
                    for col in req.dstn_only_columns:
                        val = row.get(f"{col}_dstn_only", "")
                        cell = ws.cell(row=r, column=c, value=val if val is not None else "")
                        if row.get('is_dstn_only'):
                            cell.fill = fill_dstn_only
                        else:
                            cell.fill = fill_dstn_col
                        c += 1
                
                r += 1

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Data")
        ws.append(["No records found."])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = "Comparison_Report.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return StreamingResponse(out, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == '__main__':
    # Open setup page automatically on startup
    webbrowser.open('http://127.0.0.1:8000')
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)