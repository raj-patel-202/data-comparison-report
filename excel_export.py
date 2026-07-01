import io
from typing import List, Dict, Any
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class TabData(BaseModel):
    tab_id: str
    name: str
    rows: List[Dict[str, Any]]

class ExportRequest(BaseModel):
    id_col_name: str
    columns: List[str]
    src_only_columns: List[str]
    dstn_only_columns: List[str]
    tabs: List[TabData]

def generate_excel_workbook(req: ExportRequest) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    
    fill_src_only = PatternFill(start_color="E2F0FF", end_color="E2F0FF", fill_type="solid")
    fill_dstn_only = PatternFill(start_color="F5E8FE", end_color="F5E8FE", fill_type="solid")
    
    fill_src_col = PatternFill(start_color="F7FBFF", end_color="F7FBFF", fill_type="solid")
    fill_dstn_col = PatternFill(start_color="FCF7FF", end_color="FCF7FF", fill_type="solid")
    font_src_col = Font(bold=True, color="0A84FF")
    font_dstn_col = Font(bold=True, color="BF5AF2")
    
    for tab in req.tabs:
        ws = wb.create_sheet(title=tab.name)
        
        # --- Add Legend to Top of Sheet ---
        ws.append(["Legend:"])
        ws.cell(row=1, column=1).font = Font(bold=True)
        
        ws.append(["Source Data", "data/columns (Source)"])
        ws.cell(row=2, column=1).font = font_src_col
        ws.cell(row=2, column=1).fill = fill_src_col
        
        ws.append(["Target Data", "data/columns (Target)"])
        ws.cell(row=3, column=1).font = font_dstn_col
        ws.cell(row=3, column=1).fill = fill_dstn_col
        
        ws.append(["Source Mismatch/Missing", "mismatched/missing (Source)"])
        ws.cell(row=4, column=1).fill = fill_src_only
        
        ws.append(["Target Mismatch/Missing", "mismatched/missing (Target)"])
        ws.cell(row=5, column=1).fill = fill_dstn_only
        
        ws.append([]) # Blank row 6
        
        header_row = 7
        
        if tab.tab_id == 'sub-non-comparable-columns':
            headers = [f"{req.id_col_name} (Source)"] + req.src_only_columns + [""] + [f"{req.id_col_name} (Target)"] + req.dstn_only_columns
            ws.append(headers) # This will be appended at header_row
            
            for c_idx in range(2, 2 + len(req.src_only_columns)):
                ws.cell(row=header_row, column=c_idx).font = font_src_col
                ws.cell(row=header_row, column=c_idx).fill = fill_src_col
            for c_idx in range(4 + len(req.src_only_columns), 4 + len(req.src_only_columns) + len(req.dstn_only_columns)):
                ws.cell(row=header_row, column=c_idx).font = font_dstn_col
                ws.cell(row=header_row, column=c_idx).fill = fill_dstn_col
                
            for row in tab.rows:
                row_data = [row.get('id_display', '')]
                for col in req.src_only_columns:
                    val = row.get(f"{col}_src_only", "")
                    row_data.append(val if val is not None else "")
                row_data.append("")
                row_data.append(row.get('id_display', ''))
                for col in req.dstn_only_columns:
                    val = row.get(f"{col}_dstn_only", "")
                    row_data.append(val if val is not None else "")
                ws.append(row_data)
                
                r_idx = ws.max_row
                for c_idx in range(2, 2 + len(req.src_only_columns)):
                    ws.cell(row=r_idx, column=c_idx).fill = fill_src_col
                for c_idx in range(4 + len(req.src_only_columns), 4 + len(req.src_only_columns) + len(req.dstn_only_columns)):
                    ws.cell(row=r_idx, column=c_idx).fill = fill_dstn_col
        else:
            col_idx = 1
            ws.cell(row=header_row, column=col_idx, value=req.id_col_name)
            ws.cell(row=header_row, column=col_idx).font = Font(bold=True)
            col_idx += 1
            
            for col in req.columns:
                ws.cell(row=header_row, column=col_idx, value=col)
                ws.merge_cells(start_row=header_row, start_column=col_idx, end_row=header_row, end_column=col_idx+1)
                ws.cell(row=header_row, column=col_idx).alignment = Alignment(horizontal='center')
                ws.cell(row=header_row, column=col_idx).font = Font(bold=True)
                col_idx += 2
                
            if tab.tab_id in ('sub-all', 'sub-non-comparable-src'):
                for col in req.src_only_columns:
                    ws.cell(row=header_row, column=col_idx, value=col)
                    ws.cell(row=header_row, column=col_idx).font = font_src_col
                    ws.cell(row=header_row, column=col_idx).fill = fill_src_col
                    col_idx += 1
            
            if tab.tab_id in ('sub-all', 'sub-non-comparable-dstn'):
                for col in req.dstn_only_columns:
                    ws.cell(row=header_row, column=col_idx, value=col)
                    ws.cell(row=header_row, column=col_idx).font = font_dstn_col
                    ws.cell(row=header_row, column=col_idx).fill = fill_dstn_col
                    col_idx += 1
            
            r = header_row + 1
            for row in tab.rows:
                c = 1
                ws.cell(row=r, column=c, value=row.get('id_display', ''))
                c += 1
                
                for col in req.columns:
                    is_mismatch = row.get(f"{col}_mismatch", False)
                    is_nan = row.get(f"{col}_nan", False)
                    val1 = row.get(f"{col}_1")
                    val2 = row.get(f"{col}_2")
                    
                    if val1 is None: val1 = "null"
                    if val2 is None: val2 = "null"
                    
                    if row.get('is_src_only'):
                        ws.cell(row=r, column=c, value=val1)
                        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
                        ws.cell(row=r, column=c).fill = fill_src_only
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    elif row.get('is_dstn_only'):
                        ws.cell(row=r, column=c, value=val2)
                        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
                        ws.cell(row=r, column=c).fill = fill_dstn_only
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    elif is_mismatch or is_nan:
                        c1 = ws.cell(row=r, column=c, value=val1)
                        c1.fill = fill_src_only
                        c1.alignment = Alignment(horizontal='center')
                        
                        c2 = ws.cell(row=r, column=c+1, value=val2)
                        c2.fill = fill_dstn_only
                        c2.alignment = Alignment(horizontal='center')
                    else:
                        ws.cell(row=r, column=c, value=val1)
                        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
                    
                    c += 2
                
                if tab.tab_id in ('sub-all', 'sub-non-comparable-src'):
                    for col in req.src_only_columns:
                        val = row.get(f"{col}_src_only", "")
                        cell = ws.cell(row=r, column=c, value=val if val is not None else "")
                        if row.get('is_src_only'):
                            cell.fill = fill_src_only
                        else:
                            cell.fill = fill_src_col
                        c += 1
                
                if tab.tab_id in ('sub-all', 'sub-non-comparable-dstn'):
                    for col in req.dstn_only_columns:
                        val = row.get(f"{col}_dstn_only", "")
                        cell = ws.cell(row=r, column=c, value=val if val is not None else "")
                        if row.get('is_dstn_only'):
                            cell.fill = fill_dstn_only
                        else:
                            cell.fill = fill_dstn_col
                        c += 1
                
                r += 1

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Data")
        ws.append(["No records found."])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
